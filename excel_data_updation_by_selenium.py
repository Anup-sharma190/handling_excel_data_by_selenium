"""
Excel Update Automation Demo
----------------------------
Project: Excel Automation using Python (openpyxl)

Purpose:
This script demonstrates automation skills for reading, searching, and updating Excel files.
It is designed to show practical Python skills in file handling, automation logic, and clean coding.

Tools & Skills Demonstrated:
- Python programming
- openpyxl library for Excel automation
- Searching and updating Excel data dynamically
- Error handling and safe updates
- Writing reusable and modular code
- GitHub-ready project structure

Workflow:
1. Load the Excel file
2. Identify the column based on header
3. Find the row based on search value
4. Update the cell with a new value
5. Save changes safely
6. Print confirmations and logs
"""

import openpyxl

file_path = r"C:\Users\Comp10\PycharmProjects\PythonProject\chuna.xlsx"
book = openpyxl.load_workbook(file_path)
sheet = book.active

cell_position = {}  # initialize dictionary

# Find column number for "price"
for i in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=i).value == "price":
        cell_position["col"] = i

# Find row number for "Apple"
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        if sheet.cell(row=i, column=j).value == "Apple":
            cell_position["row"] = i

# Update the value safely
if "row" in cell_position and "col" in cell_position:
    sheet.cell(row=cell_position["row"], column=cell_position["col"]).value = 500
    book.save(file_path)
    print("Updated 'Apple' price to 500!")
else:
    print("Error: 'Apple' or 'price' column not found.")
